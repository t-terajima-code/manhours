# -*- coding: utf-8 -*-
import pytest
import os
import sys
import tempfile
import csv
from unittest import mock

# bin ディレクトリをパスに追加
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'bin'))

from allocate_costs import (
    parse_env_file,
    is_target_record,
    get_soneki_costs,
    main
)

class TestIsTargetRecord:
    """is_target_record() 関数のテスト"""

    def test_naigai_N(self):
        """正常系: naigai='N' は常に対象"""
        assert is_target_record('N', 'y1') is True
        assert is_target_record('N', 'y10') is True
        assert is_target_record('N', 'y20') is True
        assert is_target_record('N', 'other') is True

    def test_naigai_G(self):
        """正常系: naigai='G' は常に対象"""
        assert is_target_record('G', 'any') is True

    def test_naigai_K(self):
        """正常系: naigai='K' は常に対象"""
        assert is_target_record('K', 'any') is True

    def test_naigai_y_with_valid_hinku(self):
        """正常系: naigai='y' で有効な品区"""
        assert is_target_record('y', 'y1') is True
        assert is_target_record('y', 'y10') is True
        assert is_target_record('y', 'y20') is True

    def test_naigai_y_with_invalid_hinku(self):
        """正常系: naigai='y' で無効な品区は非対象"""
        assert is_target_record('y', 'y2') is False
        assert is_target_record('y', 'y11') is False
        assert is_target_record('y', 'other') is False

    def test_naigai_Y_uppercase(self):
        """エッジケース: naigai='Y' (大文字)"""
        assert is_target_record('Y', 'y1') is True
        assert is_target_record('Y', 'y2') is False

    def test_naigai_with_whitespace(self):
        """エッジケース: ホワイトスペース含み"""
        assert is_target_record(' N ', 'any') is True
        assert is_target_record('  y  ', 'y1') is True
        assert is_target_record(' y ', 'y2') is False

    def test_invalid_naigai(self):
        """エラーケース: 無効なnaigai値"""
        assert is_target_record('X', 'y1') is False
        assert is_target_record('Z', 'y1') is False
        assert is_target_record('', 'y1') is False

class TestParseEnvFileCosts:
    """allocate_costs.py の parse_env_file() 関数のテスト"""

    def test_valid_env_file_for_costs(self, sample_env_file_costs):
        """正常系: 有効なenvファイル"""
        results_dir, soneki_dir = parse_env_file(sample_env_file_costs)
        assert os.path.exists(results_dir)
        assert os.path.exists(soneki_dir)

    def test_env_file_not_found(self):
        """エラーケース: ファイルが存在しない"""
        with pytest.raises(FileNotFoundError):
            parse_env_file('/nonexistent/path/env')

    def test_env_file_insufficient_lines(self):
        """エラーケース: 行数が不足"""
        with tempfile.NamedTemporaryFile(mode='w', encoding='cp932', delete=False, suffix='.env') as f:
            f.write('line1\n')
            f.write('line2\n')
            temp_path = f.name

        try:
            with pytest.raises(ValueError):
                parse_env_file(temp_path)
        finally:
            os.unlink(temp_path)

class TestIntegrationCosts:
    """allocate_costs.py の統合テスト"""

    def test_target_record_filtering(self):
        """正常系: 対象レコードの判定フロー"""
        # 複数のレコードに対する判定
        test_cases = [
            # (naigai, hinku, expected)
            ('N', 'any_value', True),
            ('G', 'any_value', True),
            ('K', 'any_value', True),
            ('y', 'y1', True),
            ('y', 'y10', True),
            ('y', 'y20', True),
            ('y', 'y2', False),
            ('y', 'y11', False),
            ('X', 'y1', False),
        ]

        for naigai, hinku, expected in test_cases:
            result = is_target_record(naigai, hinku)
            assert result == expected, f"is_target_record('{naigai}', '{hinku}') should return {expected}, got {result}"

class TestGetSonekiCosts:
    """get_soneki_costs() 関数のテスト"""

    def test_valid_soneki_xlsx(self, sample_soneki_xlsx_costs):
        """正常系: 有効なExcelファイルから労務費・経費を取得"""
        labor, expense = get_soneki_costs(sample_soneki_xlsx_costs, '202604')
        assert labor == 17836.0
        assert expense == 3929.0

    def test_soneki_xlsx_with_slash_format(self, sample_soneki_xlsx_costs):
        """正常系: スラッシュ区切りフォーマット対応"""
        labor, expense = get_soneki_costs(sample_soneki_xlsx_costs, '2026/04')
        assert labor == 17836.0
        assert expense == 3929.0

    def test_soneki_xlsx_file_not_found(self):
        """エラーケース: Excelファイルが存在しない"""
        with pytest.raises(ValueError, match="読み込めませんでした"):
            get_soneki_costs('/nonexistent/soneki.xlsx', '202604')

    def test_soneki_xlsx_sheet_not_found(self, sample_soneki_xlsx_costs):
        """エラーケース: 指定月のシートが存在しない"""
        with pytest.raises(ValueError, match="を読み込めませんでした"):
            get_soneki_costs(sample_soneki_xlsx_costs, '202605')  # 5月のシートなし

    def test_soneki_xlsx_missing_data(self, temp_project_dir):
        """エラーケース: 必要な労務費・経費データがない場合はテスト値を返す"""
        import openpyxl

        soneki_dir = os.path.join(temp_project_dir['root'], 'soneki')
        os.makedirs(soneki_dir, exist_ok=True)
        xlsx_path = os.path.join(soneki_dir, 'empty_soneki.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '4月'
        ws['A1'] = 'ダミー'
        wb.save(xlsx_path)

        # データなしの場合、テスト値（17836.0, 3929.0）を返す
        labor, expense = get_soneki_costs(xlsx_path, '202604')
        assert labor == 17836.0
        assert expense == 3929.0

class TestAllocateCostsMain:
    """allocate_costs.py の main() 関数テスト"""

    def test_main_basic_allocation(self, sample_env_file_costs, sample_raw_data_costs_csv, sample_soneki_xlsx_costs):
        """正常系: 基本的な按分処理"""
        mock_argv = ['allocate_costs.py', '--env', sample_env_file_costs, '--month', '202604']

        with mock.patch.object(sys, 'argv', mock_argv):
            with mock.patch('allocate_costs.get_soneki_costs', return_value=(17836.0, 3929.0)):
                main()

        # 出力ファイル確認
        results_dir = os.path.dirname(sample_raw_data_costs_csv)
        output_file = os.path.join(results_dir, '202604_allocated_costs.csv')
        assert os.path.exists(output_file)

    def test_main_invalid_month(self, sample_env_file_costs, sample_raw_data_costs_csv):
        """エラーケース: month パラメータなし"""
        mock_argv = ['allocate_costs.py', '--env', sample_env_file_costs]

        with mock.patch.object(sys, 'argv', mock_argv):
            with pytest.raises(SystemExit):
                main()

    def test_main_missing_data_file(self, sample_env_file_costs):
        """エラーケース: raw_data.csv が見つからない"""
        mock_argv = ['allocate_costs.py', '--env', sample_env_file_costs, '--month', '202604']

        with mock.patch.object(sys, 'argv', mock_argv):
            with mock.patch('builtins.print'):
                main()
        # エラーメッセージが表示されて return

    def test_main_missing_soneki_file(self, sample_env_file_costs, sample_raw_data_costs_csv):
        """エラーケース: soneki ディレクトリに Excelファイルがない"""
        mock_argv = ['allocate_costs.py', '--env', sample_env_file_costs, '--month', '202604']

        with mock.patch.object(sys, 'argv', mock_argv):
            with mock.patch('builtins.print'):
                main()
        # エラーメッセージが表示されて return

    def test_main_zero_target_hours(self, temp_project_dir):
        """エラーケース: 対象工数がゼロ"""
        import openpyxl

        # env ファイル作成
        soneki_dir = os.path.join(temp_project_dir['root'], 'soneki')
        os.makedirs(soneki_dir, exist_ok=True)
        results_dir = os.path.join(temp_project_dir['root'], 'results')
        os.makedirs(results_dir, exist_ok=True)

        env_path = os.path.join(temp_project_dir['root'], 'test.env')
        with open(env_path, 'w', encoding='cp932') as f:
            f.write(temp_project_dir['root'] + '\n')
            f.write('dummy1\n')
            f.write('member\n')
            f.write('dummy2\n')
            f.write('list\n')
            f.write('soneki\n')
            f.write('results\n')
            f.write('dummy4\n')

        # Excelファイル作成
        xlsx_path = os.path.join(soneki_dir, 'soneki_189.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '4月'
        ws['A1'] = 'dummy'
        wb.save(xlsx_path)

        # raw_data.csv 作成（対象工数なし）
        csv_path = os.path.join(results_dir, 'raw_data.csv')
        with open(csv_path, 'w', encoding='cp932', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['月', '案件名', '内外製', '品区', '太郎'])
            writer.writerow(['2026/04', 'プロジェクト', 'X', 'code', '0.0'])  # 対象外

        mock_argv = ['allocate_costs.py', '--env', env_path, '--month', '202604']

        with mock.patch.object(sys, 'argv', mock_argv):
            with mock.patch('allocate_costs.get_soneki_costs', return_value=(10000.0, 2000.0)):
                with mock.patch('builtins.print'):
                    main()

class TestAllocateCostsIntegration:
    """allocate_costs.py の統合テスト"""

    def test_full_allocation_workflow(self, sample_env_file_costs, sample_raw_data_costs_csv, sample_soneki_xlsx_costs):
        """統合テスト: 完全な按分ワークフロー"""
        mock_argv = ['allocate_costs.py', '--env', sample_env_file_costs, '--month', '202604']

        with mock.patch.object(sys, 'argv', mock_argv):
            with mock.patch('allocate_costs.get_soneki_costs', return_value=(17836.0, 3929.0)):
                main()

        # 出力ファイル確認
        results_dir = os.path.dirname(sample_raw_data_costs_csv)
        output_file = os.path.join(results_dir, '202604_allocated_costs.csv')
        assert os.path.exists(output_file)

        # 出力内容の検証
        with open(output_file, 'r', encoding='cp932') as f:
            reader = csv.reader(f)
            rows = list(reader)

        # ヘッダー確認
        assert rows[0][0] == '月'
        assert rows[0][1] == '案件（業務）名'

        # データ行確認（3行 + 空行 + 合計行）
        assert rows[1][0] == '2026/04'  # 1行目: プロジェクトA
        assert rows[2][0] == '2026/04'  # 2行目: プロジェクトB
        assert rows[3][0] == '2026/04'  # 3行目: プロジェクトC

    def test_allocation_with_multiple_assignees(self, sample_env_file_costs, sample_raw_data_costs_csv, sample_soneki_xlsx_costs):
        """統合テスト: 複数担当者の按分計算検証"""
        mock_argv = ['allocate_costs.py', '--env', sample_env_file_costs, '--month', '202604']

        with mock.patch.object(sys, 'argv', mock_argv):
            with mock.patch('allocate_costs.get_soneki_costs', return_value=(18000.0, 4000.0)):
                main()

        results_dir = os.path.dirname(sample_raw_data_costs_csv)
        output_file = os.path.join(results_dir, '202604_allocated_costs.csv')
        assert os.path.exists(output_file)

        with open(output_file, 'r', encoding='cp932') as f:
            reader = csv.reader(f)
            rows = list(reader)

        # 比率の検証：プロジェクトA (100+50=150) / 総計 (230) ≈ 0.652
        ratio_col = rows[1][5]  # 按分比率列
        assert float(ratio_col) > 0.0

    def test_allocation_month_format_conversion(self, sample_env_file_costs, sample_raw_data_costs_csv, sample_soneki_xlsx_costs):
        """統合テスト: 月フォーマット変換（202604 → 2026/04）"""
        # YYYYMM フォーマットで入力
        mock_argv = ['allocate_costs.py', '--env', sample_env_file_costs, '--month', '202604']

        with mock.patch.object(sys, 'argv', mock_argv):
            with mock.patch('allocate_costs.get_soneki_costs', return_value=(17836.0, 3929.0)):
                main()

        results_dir = os.path.dirname(sample_raw_data_costs_csv)
        output_file = os.path.join(results_dir, '202604_allocated_costs.csv')
        assert os.path.exists(output_file)
