# -*- coding: utf-8 -*-
import pytest
import os
import tempfile
import csv
from pathlib import Path

@pytest.fixture
def temp_project_dir():
    """テスト用の一時的なプロジェクトディレクトリを作成"""
    with tempfile.TemporaryDirectory() as tmpdir:
        # ディレクトリ構造を作成
        member_dir = os.path.join(tmpdir, 'member')
        list_dir = os.path.join(tmpdir, 'list')
        results_dir = os.path.join(tmpdir, 'results')

        os.makedirs(member_dir)
        os.makedirs(list_dir)
        os.makedirs(results_dir)

        yield {
            'root': tmpdir,
            'member_dir': member_dir,
            'list_dir': list_dir,
            'results_dir': results_dir
        }

@pytest.fixture
def sample_env_file(temp_project_dir):
    """テスト用のenvファイルを作成"""
    env_path = os.path.join(temp_project_dir['root'], 'test.env')
    with open(env_path, 'w', encoding='cp932') as f:
        f.write(temp_project_dir['root'] + '\n')
        f.write('dummy1\n')
        f.write('member\n')
        f.write('dummy2\n')
        f.write('list\n')
        f.write('dummy3\n')
        f.write('results\n')
        f.write('dummy4\n')
    return env_path

@pytest.fixture
def sample_nichijou_csv(temp_project_dir):
    """テスト用の日常作業マスタCSVを作成"""
    csv_path = os.path.join(temp_project_dir['list_dir'], 'test_nichijou_list.csv')
    with open(csv_path, 'w', encoding='cp932', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['部課', '工数区分', 'タスク名'])
        writer.writerow(['営業部', 'A', 'テスト_001'])
        writer.writerow(['営業部', 'B', 'テスト_002'])
    return csv_path

@pytest.fixture
def sample_proj_csv(temp_project_dir):
    """テスト用のプロジェクトマスタCSVを作成"""
    csv_path = os.path.join(temp_project_dir['list_dir'], 'test_proj_list.csv')
    with open(csv_path, 'w', encoding='cp932', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['プロジェクトコード', '部課', '工数区分', 'プロジェクト名'])
        writer.writerow(['PRJ001', '営業部', 'A', 'プロジェクトA'])
        writer.writerow(['PRJ002', '営業部', 'B', 'プロジェクトB'])
    return csv_path

@pytest.fixture
def sample_staff_csv(temp_project_dir):
    """テスト用の担当者CSVを作成 (44000=2020/06 なので 202006_ プレフィックス)"""
    csv_path = os.path.join(temp_project_dir['member_dir'], '202006_T_test_user.csv')
    with open(csv_path, 'w', encoding='cp932', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['日付', 'タスク名', '工数'])
        writer.writerow(['44000', 'テスト_001', '8'])
        writer.writerow(['44001', 'テスト_002', '8'])
    return csv_path

@pytest.fixture
def sample_env_file_costs(temp_project_dir):
    """allocate_costs.py用のテスト用envファイルを作成"""
    soneki_dir = os.path.join(temp_project_dir['root'], 'soneki')
    os.makedirs(soneki_dir, exist_ok=True)

    env_path = os.path.join(temp_project_dir['root'], 'test_costs.env')
    with open(env_path, 'w', encoding='cp932') as f:
        f.write(temp_project_dir['root'] + '\n')
        f.write('dummy1\n')
        f.write('member\n')
        f.write('dummy2\n')
        f.write('list\n')
        f.write('soneki\n')
        f.write('results\n')
        f.write('dummy4\n')
    return env_path

@pytest.fixture
def sample_env_file_actual_hours(temp_project_dir):
    """allocate_actual_hours.py用のテスト用envファイルを作成（標準時間有り）"""
    kintai_dir = os.path.join(temp_project_dir['root'], 'kintai')
    os.makedirs(kintai_dir, exist_ok=True)

    env_path = os.path.join(temp_project_dir['root'], 'test_actual_hours.env')
    with open(env_path, 'w', encoding='cp932') as f:
        f.write(temp_project_dir['root'] + '\n')
        f.write('dummy1\n')
        f.write('member\n')
        f.write('kintai\n')
        f.write('list\n')
        f.write('dummy3\n')
        f.write('results\n')
        f.write('8.0\n')
    return env_path

@pytest.fixture
def sample_env_file_actual_hours_no_time(temp_project_dir):
    """allocate_actual_hours.py用のテスト用envファイルを作成（標準時間なし）"""
    kintai_dir = os.path.join(temp_project_dir['root'], 'kintai')
    os.makedirs(kintai_dir, exist_ok=True)

    env_path = os.path.join(temp_project_dir['root'], 'test_actual_hours_notime.env')
    with open(env_path, 'w', encoding='cp932') as f:
        f.write(temp_project_dir['root'] + '\n')
        f.write('dummy1\n')
        f.write('member\n')
        f.write('kintai\n')
        f.write('list\n')
        f.write('dummy3\n')
        f.write('results\n')
        f.write('invalid_value\n')
    return env_path

@pytest.fixture
def sample_env_file_verify_hours(temp_project_dir):
    """verify_hours.py用のテスト用envファイルを作成"""
    kintai_dir = os.path.join(temp_project_dir['root'], 'kintai')
    os.makedirs(kintai_dir, exist_ok=True)

    env_path = os.path.join(temp_project_dir['root'], 'test_verify_hours.env')
    with open(env_path, 'w', encoding='cp932') as f:
        f.write(temp_project_dir['root'] + '\n')
        f.write('dummy1\n')
        f.write('member\n')
        f.write('kintai\n')
        f.write('list\n')
        f.write('dummy3\n')
        f.write('results\n')
        f.write('dummy4\n')
    return env_path

@pytest.fixture
def sample_kintai_html(temp_project_dir):
    """テスト用の勤怠HTMLを作成"""
    html_path = os.path.join(temp_project_dir['root'], 'test_kintai.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write('''<html>
<body>
<table>
<tr><th>名前</th><th>雇用区分</th><th>実働時間</th></tr>
<tr><td>太郎</td><td>正社員</td><td>8.0</td></tr>
<tr><td>花子</td><td>契約社員</td><td>7.5</td></tr>
</table>
</body>
</html>''')
    return html_path

@pytest.fixture
def sample_kintai_html_with_numbers(temp_project_dir):
    """テスト用の勤怠HTML（先頭に番号あり）を作成"""
    html_path = os.path.join(temp_project_dir['root'], 'test_kintai_numbers.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write('''<html>
<body>
<table>
<tr><th>名前</th><th>雇用区分</th><th>実働時間</th></tr>
<tr><td>001太郎</td><td>正社員</td><td>8.0</td></tr>
</table>
</body>
</html>''')
    return html_path

@pytest.fixture
def sample_kintai_html_with_spaces(temp_project_dir):
    """テスト用の勤怠HTML（スペース付き）を作成"""
    html_path = os.path.join(temp_project_dir['root'], 'test_kintai_spaces.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write('''<html>
<body>
<table>
<tr><th>名前</th><th>雇用区分</th><th>実働時間</th></tr>
<tr><td>太　郎</td><td>正社員</td><td>8.0</td></tr>
</table>
</body>
</html>''')
    return html_path

@pytest.fixture
def sample_kintai_html_missing_headers(temp_project_dir):
    """テスト用の勤怠HTML（ヘッダーなし）を作成"""
    html_path = os.path.join(temp_project_dir['root'], 'test_kintai_no_headers.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write('''<html>
<body>
<table>
<tr><th>従業員</th><th>身分</th><th>勤務</th></tr>
<tr><td>太郎</td><td>正社員</td><td>8.0</td></tr>
</table>
</body>
</html>''')
    return html_path

@pytest.fixture
def sample_env_file_person_months(temp_project_dir):
    """calculate_person_months.py用のテスト用envファイルを作成"""
    env_path = os.path.join(temp_project_dir['root'], 'test_person_months.env')
    with open(env_path, 'w', encoding='cp932') as f:
        f.write(temp_project_dir['root'] + '\n')
        f.write('dummy1\n')
        f.write('member\n')
        f.write('dummy2\n')
        f.write('list\n')
        f.write('dummy3\n')
        f.write('results\n')
        f.write('dummy4\n')
    return env_path

@pytest.fixture
def sample_raw_data_csv(temp_project_dir):
    """テスト用のraw_data.csvを作成"""
    csv_path = os.path.join(temp_project_dir['results_dir'], 'raw_data.csv')
    with open(csv_path, 'w', encoding='cp932', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['月', '案件（業務）名', '内外製区分', '品区コード', '太郎', '花子'])
        writer.writerow(['2026/05', 'プロジェクトA', 'N', 'code1', '8.0', '8.0'])
        writer.writerow(['2026/05', 'プロジェクトB', 'G', 'code2', '8.0', '8.0'])
    return csv_path

@pytest.fixture
def sample_env_file_export_keiri(temp_project_dir):
    """export_keiri_csv.py用のテスト用envファイルを作成"""
    env_path = os.path.join(temp_project_dir['root'], 'test_export_keiri.env')
    with open(env_path, 'w', encoding='cp932') as f:
        f.write(temp_project_dir['root'] + '\n')
        f.write('dummy1\n')
        f.write('member\n')
        f.write('dummy2\n')
        f.write('list\n')
        f.write('dummy3\n')
        f.write('results\n')
        f.write('dummy4\n')
    return env_path

@pytest.fixture
def sample_hinku_list_csv(temp_project_dir):
    """テスト用のhinku_list.csvを作成"""
    csv_path = os.path.join(temp_project_dir['list_dir'], '189_hinku_list.csv')
    with open(csv_path, 'w', encoding='cp932', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['内外製区分', '品区コード'])
        writer.writerow(['N', 'code1'])
        writer.writerow(['G', 'code2'])
        writer.writerow(['y', 'y2'])
    return csv_path

@pytest.fixture
def sample_env_file_update_data(temp_project_dir):
    """update_data.py用のテスト用envファイルを作成"""
    env_path = os.path.join(temp_project_dir['root'], 'test_update_data.env')
    with open(env_path, 'w', encoding='utf-8') as f:
        f.write(temp_project_dir['root'] + '\n')
        f.write('dummy1\n')
        f.write('member\n')
        f.write('dummy2\n')
        f.write('list\n')
        f.write('dummy3\n')
        f.write('results\n')
        f.write('dummy4\n')
    return env_path

@pytest.fixture
def sample_raw_data_costs_csv(temp_project_dir):
    """allocate_costs.py用のraw_data.csvを作成"""
    csv_path = os.path.join(temp_project_dir['results_dir'], 'raw_data.csv')
    with open(csv_path, 'w', encoding='cp932', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['月', '案件（業務）名', '内外製区分', '品区コード', '太郎', '花子'])
        writer.writerow(['2026/04', 'プロジェクトA', 'N', 'code1', '100.0', '50.0'])
        writer.writerow(['2026/04', 'プロジェクトB', 'G', 'code2', '80.0', '120.0'])
        writer.writerow(['2026/04', 'プロジェクトC', 'y', 'y1', '50.0', '30.0'])
        writer.writerow(['2026/05', 'プロジェクトA', 'N', 'code1', '75.0', '65.0'])
    return csv_path

@pytest.fixture
def sample_soneki_xlsx_costs(temp_project_dir):
    """allocate_costs.py用のsoneki Excelファイルを作成"""
    import openpyxl

    soneki_dir = os.path.join(temp_project_dir['root'], 'soneki')
    os.makedirs(soneki_dir, exist_ok=True)
    xlsx_path = os.path.join(soneki_dir, 'soneki_189.xlsx')

    # Excelワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '4月'

    # ダミーヘッダーと「全社」「実績金額」ヘッダー
    ws['A1'] = 'カテゴリ1'
    ws['B1'] = 'カテゴリ2'
    ws['C1'] = 'カテゴリ3'
    ws['D1'] = '全社'
    ws['E1'] = '実績金額'

    # (14) 機能部間接部門費（労務費）セクション
    ws['A3'] = '(14) 機能部間接部門費（労務費）'
    ws['A4'] = '研究開発'
    ws['E4'] = '17836'  # 労務費

    # (15) 機能部間接部門費（経費）セクション
    ws['A6'] = '(15) 機能部間接部門費（経費）'
    ws['A7'] = '研究開発'
    ws['E7'] = '3929'  # 経費

    wb.save(xlsx_path)
    return xlsx_path
