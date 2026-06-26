# -*- coding: utf-8 -*-
"""
generate_dashboard_data.py
results/ フォルダの CSV を統合して dashboard/data/data.js を生成する

実行方法:
    cd manhours/bin
    python generate_dashboard_data.py
"""
import os
import glob
import json
import math
import re
import csv as _csv
from datetime import datetime
import pandas as pd

_MONTH_RE = re.compile(r'^\d{4}/\d{2}$')

FIXED_HOUR_KEYS = {"月", "案件（業務）名", "区分", "内外製区分", "品区コード", "合計時間(h)"}

# ダッシュボードに表示しないグループ（日常業務・教育・不働工数）
EXCLUDE_GROUPS = {"日常業務", "教育", "不働工数"}

# グループ → 5区分マッピング
KUBUN_MAP = {
    "新製品PJ": "開発テーマ",
    "MFS":      "開発テーマ",
    "精練":     "開発テーマ",
    "FH":       "開発テーマ",
    "工用":     "インシデント",
    "電気":     "インシデント",
    "ロール":   "インシデント",
    "合樹":     "インシデント",
    "印材":     "インシデント",
    "(その他)": "インシデント",
    "日常業務": "日常業務",
    "教育":     "教育",
    "不働工数": "不働工数",
}


def load_standard_time(bin_dir):
    """bin/env ファイルの8行目から標準時間（h/日）を読み取る"""
    env_path = os.path.join(bin_dir, 'env')
    try:
        with open(env_path, 'r', encoding='cp932') as f:
            lines = [l.rstrip('\n') for l in f.readlines()]
        # 8行目（0-indexed: 7行目）
        if len(lines) >= 8:
            return float(lines[7].strip())
    except Exception as e:
        print(f"  警告: env ファイル読み込みエラー: {e}")
    return 7.5  # デフォルト


def compute_fiscal_year(month_str):
    """'YYYY/MM' → 事業年度番号 (188, 189, ...)
    4月始まり: 2025/04〜2026/03 = 188期, 2026/04〜2027/03 = 189期"""
    yyyy, mm = month_str.split('/')
    return int(yyyy) - (1837 if int(mm) >= 4 else 1838)


def load_period_proj_list(list_dir):
    """NNN_proj_list.csv を期番号別に読み込む → {期: {proj_name: jigyou}}"""
    period_map = {}
    for path in sorted(glob.glob(os.path.join(list_dir, '*_proj_list.csv'))):
        basename = os.path.basename(path)
        m = re.match(r'^(\d{3})_', basename)
        if not m:
            continue
        period = int(m.group(1))
        try:
            df = pd.read_csv(path, encoding='cp932')
            mapping = {}
            for _, row in df.iterrows():
                name   = str(row.get('proj_name', '')).strip()
                jigyou = str(row.get('jigyou',    '')).strip()
                if name and jigyou and name != 'nan':
                    mapping[name] = jigyou
            period_map[period] = mapping
        except Exception as e:
            print(f"  警告: {basename} 読み込みエラー: {e}")
    return period_map


def load_period_inc_list(list_dir):
    """NNN_inc_list.csv を期番号別に読み込む → {期: {inc_num: (naigaikubun, hinku)}}"""
    period_map = {}
    for path in sorted(glob.glob(os.path.join(list_dir, '*_inc_list.csv'))):
        basename = os.path.basename(path)
        m = re.match(r'^(\d{3})_', basename)
        if not m:
            continue
        period = int(m.group(1))
        try:
            df = pd.read_csv(path, encoding='cp932')
            mapping = {}
            for _, row in df.iterrows():
                inc_num = str(row.get('inc_num',      '')).strip()
                naigai  = str(row.get('naigaikubun',  '')).strip()
                hinku   = str(row.get('hinku',        '')).strip()
                if inc_num and inc_num != 'nan':
                    mapping[inc_num] = (naigai, hinku)
            period_map[period] = mapping
        except Exception as e:
            print(f"  警告: {basename} 読み込みエラー: {e}")
    return period_map


def load_hinku_list(list_dir):
    """*_hinku_list.csv / hinku_list.csv を読み込み、item マッピングと group マッピングを返す"""
    item_map  = {}
    group_map = {}
    # 優先順位: 番号付き (188_, 189_, ...) → 旧来の hinku_list.csv
    paths = sorted(glob.glob(os.path.join(list_dir, '*_hinku_list.csv')))
    if not paths:
        legacy = os.path.join(list_dir, 'hinku_list.csv')
        if os.path.exists(legacy):
            paths = [legacy]
    if not paths:
        return item_map, group_map
    for path in paths:
        try:
            df = pd.read_csv(path, encoding='cp932')
            for _, row in df.iterrows():
                naigai = str(row.get('naigaikubun', '')).strip()
                hinku  = str(row.get('hinku',       '')).strip()
                item   = str(row.get('item',         '')).strip()
                group  = str(row.get('group',        '')).strip()
                if not (naigai and hinku):
                    continue
                key = (naigai, hinku)
                if item and item != 'nan':
                    item_map[key] = item
                if group and group != 'nan':
                    group_map[key] = group
        except Exception as e:
            print(f"  警告: {os.path.basename(path)} 読み込みエラー: {e}")
    return item_map, group_map


def get_group(proj_name, naigaikubun, hinku, month, period_proj, period_inc, hinku_group_map):
    """月から期を決定し、期別の proj_list(開発テーマ) / inc_list(インシデント) でグループを判定する"""
    try:
        period = compute_fiscal_year(month)
    except Exception:
        period = None

    # 1. 期別 proj_list で開発テーマか確認
    if period is not None:
        proj_map = period_proj.get(period, {})
        if proj_name in proj_map:
            return proj_map[proj_name]

    # 2. 期別 inc_list でインシデントか確認（inc_numが一致→hinku_listでサブグループ）
    if period is not None:
        inc_map = period_inc.get(period, {})
        if proj_name in inc_map:
            inc_naigai, inc_hinku = inc_map[proj_name]
            key = (inc_naigai, inc_hinku)
            if key in hinku_group_map:
                return hinku_group_map[key]

    # 3. hinku_list フォールバック（x系の日常業務/教育/不働工数 など）
    key = (str(naigaikubun).strip(), str(hinku).strip())
    if key in hinku_group_map:
        return hinku_group_map[key]

    return "(その他)"


def get_kubun(proj_name, naigaikubun, hinku, month, period_proj, period_inc, hinku_group_map):
    """kubunHours用: 月→期を決定し5区分kubunを返す。
    proj_listに含まれる案件は jigyou に関わらず常に「開発テーマ」と判定する。"""
    try:
        period = compute_fiscal_year(month)
    except Exception:
        period = None

    # 1. proj_listに含まれる → 常に開発テーマ
    if period is not None:
        if proj_name in period_proj.get(period, {}):
            return "開発テーマ"

    # 2. inc_listに含まれる → hinku_listでサブグループ → KUBUN_MAP
    if period is not None:
        inc_map = period_inc.get(period, {})
        if proj_name in inc_map:
            inc_naigai, inc_hinku = inc_map[proj_name]
            group = hinku_group_map.get((inc_naigai, inc_hinku), "(その他)")
            return KUBUN_MAP.get(group, "インシデント")

    # 3. hinku_listフォールバック（x系の日常業務/教育/不働工数など）
    group = hinku_group_map.get((str(naigaikubun).strip(), str(hinku).strip()), "(その他)")
    return KUBUN_MAP.get(group, "インシデント")


def get_item(naigaikubun, hinku, item_map):
    """内外製区分と品区コードからアイテム名を決定する"""
    key = (str(naigaikubun).strip(), str(hinku).strip())
    return item_map.get(key, str(hinku))


def sanitize(v):
    """JSON シリアライズ用に値を正規化する"""
    if v is None:
        return 0
    try:
        if isinstance(v, float):
            if math.isnan(v) or math.isinf(v):
                return 0.0
            return round(v, 6)
        if pd.isna(v):
            return 0
    except Exception:
        pass
    return v


def load_all_costs(results_dir, period_proj, period_inc, item_map, hinku_group_map):
    """全月の allocated_costs.csv を統合してレコードリストを返す"""
    pattern = os.path.join(results_dir, '??????_allocated_costs.csv')
    files = sorted(glob.glob(pattern))
    records = []
    skipped = 0
    for fpath in files:
        try:
            df = pd.read_csv(fpath, encoding='cp932')
            for _, row in df.iterrows():
                proj_name = str(row.get('案件（業務）名', '')).strip()
                naigai    = str(row.get('内外製区分',     '')).strip()
                hinku     = str(row.get('品区コード',     '')).strip()
                month     = str(row.get('月',            '')).strip()
                if (proj_name in ('', 'nan') or naigai == 'nan'
                        or not _MONTH_RE.match(month)
                        or '.xlsx' in proj_name):
                    continue
                group = get_group(proj_name, naigai, hinku, month, period_proj, period_inc, hinku_group_map)
                if group in EXCLUDE_GROUPS:
                    skipped += 1
                    continue
                records.append({
                    "月": month,
                    "案件（業務）名": proj_name,
                    "内外製区分": naigai,
                    "品区コード": hinku,
                    "対象工数(hour)": sanitize(row.get('対象工数(hour)', 0)),
                    "按分比率":       sanitize(row.get('按分比率',       0)),
                    "投入人員(人)":   sanitize(row.get('投入人員(人)',   0)),
                    "労務費(千円)":   sanitize(row.get('労務費(千円)',   0)),
                    "経費(千円)":     sanitize(row.get('経費(千円)',     0)),
                    "group": group,
                    "item":  get_item(naigai, hinku, item_map),
                })
        except Exception as e:
            print(f"  警告: {os.path.basename(fpath)} 読み込みエラー: {e}")
    if skipped:
        print(f"  除外グループ（日常業務・教育・不働工数）: {skipped} 件スキップ")
    return records


def load_all_hours(results_dir, period_proj, period_inc, item_map, hinku_group_map):
    """全月の allocated_hours.csv を統合してレコードリストを返す（メンバー列含む）"""
    pattern = os.path.join(results_dir, '??????_allocated_hours.csv')
    files = sorted(glob.glob(pattern))

    all_dfs = []
    for fpath in files:
        try:
            df = pd.read_csv(fpath, encoding='cp932')
            all_dfs.append(df)
        except Exception as e:
            print(f"  警告: {os.path.basename(fpath)} 読み込みエラー: {e}")

    if not all_dfs:
        return []

    all_cols: set = set()
    for df in all_dfs:
        all_cols.update(df.columns.tolist())
    member_cols = sorted(c for c in all_cols if c not in FIXED_HOUR_KEYS)

    records = []
    skipped = 0
    for df in all_dfs:
        for mc in member_cols:
            if mc not in df.columns:
                df[mc] = 0.0
            else:
                df[mc] = pd.to_numeric(df[mc], errors='coerce').fillna(0.0)
        df[member_cols] = df[member_cols].fillna(0.0)

        for _, row in df.iterrows():
            proj_name = str(row.get('案件（業務）名', '')).strip()
            naigai    = str(row.get('内外製区分',     '')).strip()
            hinku     = str(row.get('品区コード',     '')).strip()
            month     = str(row.get('月',            '')).strip()
            if (proj_name in ('', 'nan') or naigai == 'nan'
                    or not _MONTH_RE.match(month)
                    or '.xlsx' in proj_name):
                continue
            group = get_group(proj_name, naigai, hinku, month, period_proj, period_inc, hinku_group_map)
            if group in EXCLUDE_GROUPS:
                skipped += 1
                continue
            record = {
                "月": month,
                "案件（業務）名": proj_name,
                "区分":         str(row.get('区分', '')).strip(),
                "内外製区分":   naigai,
                "品区コード":   hinku,
                "合計時間(h)":  sanitize(row.get('合計時間(h)', 0)),
                "group": group,
                "item":  get_item(naigai, hinku, item_map),
            }
            for mc in member_cols:
                record[mc] = sanitize(row.get(mc, 0))
            records.append(record)

    if skipped:
        print(f"  除外グループ（日常業務・教育・不働工数）: {skipped} 件スキップ")
    return records


def load_kubun_hours(results_dir, period_proj, period_inc, hinku_group_map, standard_time):
    """全月の allocated_hours.csv から 5区分別・月別累計工数を返す（除外グループ・休暇含む）"""
    pattern = os.path.join(results_dir, '??????_allocated_hours.csv')
    files = sorted(glob.glob(pattern))
    # {月: {kubun: hours}}
    agg = {}
    kyuka_total = 0
    for fpath in files:
        try:
            df = pd.read_csv(fpath, encoding='cp932')
            for _, row in df.iterrows():
                month     = str(row.get('月', '')).strip()
                proj_name = str(row.get('案件（業務）名', '')).strip()
                if not _MONTH_RE.match(month):
                    continue
                if proj_name in ('', 'nan') or '.xlsx' in proj_name:
                    continue

                hours = pd.to_numeric(row.get('合計時間(h)', 0), errors='coerce')
                if isinstance(hours, float) and math.isnan(hours):
                    hours = 0.0

                # 休暇行（内外製区分=NaN）: 合計時間(h) = 休暇日数合計 × 標準時間 → 不働工数へ
                if proj_name == '休暇':
                    if month not in agg:
                        agg[month] = {}
                    agg[month]['不働工数'] = agg[month].get('不働工数', 0.0) + float(hours)
                    kyuka_total += float(hours)
                    continue

                naigai = str(row.get('内外製区分', '')).strip()
                hinku  = str(row.get('品区コード', '')).strip()
                if naigai == 'nan':
                    continue

                kubun = get_kubun(proj_name, naigai, hinku, month, period_proj, period_inc, hinku_group_map)
                if month not in agg:
                    agg[month] = {}
                agg[month][kubun] = agg[month].get(kubun, 0.0) + float(hours)
        except Exception as e:
            print(f"  警告: {os.path.basename(fpath)} 読み込みエラー: {e}")
    if kyuka_total > 0:
        print(f"  休暇工数（標準時間={standard_time}h/日）: {round(kyuka_total)} h → 不働工数へ加算")

    records = []
    for month in sorted(agg):
        for kubun, hours in agg[month].items():
            records.append({"月": month, "kubun": kubun, "hours": round(hours, 2)})
    return records


def load_person_months(results_dir):
    """全月の person_months.csv から月別合計人工数を返す"""
    pattern = os.path.join(results_dir, '??????_person_months.csv')
    files = sorted(glob.glob(pattern))
    records = []
    for fpath in files:
        try:
            df = pd.read_csv(fpath, encoding='cp932')
            month_val = None
            month_total = 0.0
            for _, row in df.iterrows():
                m = str(row.get('月', '')).strip()
                if not _MONTH_RE.match(m):
                    continue
                if month_val is None:
                    month_val = m
                pm = pd.to_numeric(row.get('人工数/月', 0), errors='coerce')
                if not (isinstance(pm, float) and math.isnan(pm)):
                    month_total += float(pm)
            if month_val and month_total > 0:
                records.append({"月": month_val, "pm": round(month_total, 4)})
        except Exception as e:
            print(f"  警告: {os.path.basename(fpath)} 読み込みエラー: {e}")
    return records


_PERIOD_SHEET_RE = re.compile(r'◆\d+期$')

def load_inc_name_map(list_dir, xlsx_path=None):
    """インシデント番号 → 件名 のマッピングを返す。

    まず list/inc_name_list.csv（配布物に同梱・マスタから生成済み, CP932,
    ヘッダー inc_num,name）を読む。配布環境ではマスタ xlsx を持たないため、
    この CSV があれば xlsx 無しでも案件名を復元できる。
    CSV が無い／空の場合のみ インシデント管理表.xlsx（開発環境のマスタ）に
    フォールバックする。
    """
    name_map = {}
    csv_path = os.path.join(list_dir, 'inc_name_list.csv')
    if os.path.exists(csv_path):
        try:
            with open(csv_path, 'r', encoding='cp932', newline='') as f:
                reader = _csv.reader(f)
                next(reader, None)  # ヘッダー行（inc_num,name）をスキップ
                for row in reader:
                    if len(row) < 2:
                        continue
                    inc_num = row[0].strip()
                    name = row[1].strip()
                    if inc_num and name:
                        name_map[inc_num] = name
        except Exception as e:
            print(f"  警告: inc_name_list.csv 読み込みエラー: {e}")
    if name_map:
        return name_map
    # フォールバック: マスタ xlsx（開発環境のみ存在）
    if xlsx_path and os.path.exists(xlsx_path):
        return _load_inc_name_map_from_xlsx(xlsx_path)
    print("  警告: inc_name_list.csv / インシデント管理表.xlsx いずれからも"
          "案件名を取得できず、incNameMap は空になります")
    return name_map


def _load_inc_name_map_from_xlsx(xlsx_path):
    """インシデント管理表.xlsx の各期シート（◆NNN期）から
    インシデント番号 → 件名 のマッピングを返す。

    各シートのヘッダー行：
      列1: ｲﾝｼﾃﾞﾝﾄ№  列6: 件　　　名
    """
    name_map = {}
    if not os.path.exists(xlsx_path):
        print(f"  警告: {os.path.basename(xlsx_path)} が見つかりません")
        return name_map
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  警告: {os.path.basename(xlsx_path)} 読み込みエラー: {e}")
        return name_map

    for sheet_name in wb.sheetnames:
        if not _PERIOD_SHEET_RE.match(sheet_name):
            continue
        ws = wb[sheet_name]
        col_inc  = None   # インシデント番号の列インデックス
        col_name = None   # 件名の列インデックス

        for row in ws.iter_rows(values_only=True):
            if col_inc is None:
                # ヘッダー行を探す（ｲﾝｼﾃﾞﾝﾄ を含むセルを検出）
                for j, cell in enumerate(row):
                    v = str(cell or '').strip()
                    if 'ｲﾝｼﾃﾞﾝﾄ' in v:   # 半角カタカナ限定（タイトル行の全角を除外）
                        col_inc = j
                    # '件' で始まり '名' で終わるセル（'件　　　名' など）
                    stripped = v.replace('　', '').replace(' ', '')
                    if stripped.startswith('件') and stripped.endswith('名'):
                        col_name = j
            else:
                # データ行：インシデント番号を整数として取得
                if col_inc >= len(row) or col_name is None or col_name >= len(row):
                    continue
                inc_val  = row[col_inc]
                name_val = row[col_name]
                if inc_val is None or name_val is None:
                    continue
                try:
                    inc_num = str(int(inc_val))
                    name    = str(name_val).strip()
                    if inc_num and name:
                        name_map[inc_num] = name
                except (ValueError, TypeError):
                    pass

    wb.close()
    return name_map


def write_data_js(output_file, all_costs, all_hours, kubun_hours, monthly_person_months, standard_time, inc_name_map=None):
    """dashboard/data/data.js に出力する"""
    now = datetime.now().strftime('%Y/%m/%d %H:%M')
    costs_json  = json.dumps(all_costs,             ensure_ascii=False, separators=(',', ':'))
    hours_json  = json.dumps(all_hours,             ensure_ascii=False, separators=(',', ':'))
    kubun_json  = json.dumps(kubun_hours,           ensure_ascii=False, separators=(',', ':'))
    pm_json     = json.dumps(monthly_person_months, ensure_ascii=False, separators=(',', ':'))
    inc_json    = json.dumps(inc_name_map or {},    ensure_ascii=False, separators=(',', ':'))

    content = (
        f"// 自動生成ファイル - generate_dashboard_data.py で更新してください\n"
        f"// 生成日時: {now}\n"
        f"const standardTime={standard_time};\n"
        f"const allCosts={costs_json};\n"
        f"const allHours={hours_json};\n"
        f"const kubunHours={kubun_json};\n"
        f"const monthlyPersonMonths={pm_json};\n"
        f"const incNameMap={inc_json};\n"
    )
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(content)


def main():
    """エントリポイント: results/ の各CSVとマスタを統合して dashboard/data/data.js を生成する。

    bin/ の親をルートとして results/・list/・インシデント管理表.xlsx を参照し、
    allocated_costs / allocated_hours / person_months を読み込んで以下を構築する:
        allCosts            : 月別・案件別のコスト按分レコード
        allHours            : 月別・案件別の実働時間（メンバー列を含む）
        kubunHours          : 5区分別・月別の累計工数（休暇は不働工数へ加算）
        monthlyPersonMonths : 月別の合計人工数
        incNameMap          : インシデント番号→件名

    日常業務・教育・不働工数グループは allCosts/allHours から除外する。
    出力 data.js のみ UTF-8（ダッシュボードJS用）。引数は取らず env の標準時間を読む。
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir   = os.path.dirname(script_dir)

    results_dir = os.path.join(root_dir, 'results')
    list_dir    = os.path.join(root_dir, 'list')
    bin_dir     = script_dir
    inc_xlsx    = os.path.join(root_dir, 'インシデント管理表.xlsx')
    output_file = os.path.join(root_dir, 'dashboard', 'data', 'data.js')

    standard_time = load_standard_time(bin_dir)
    print(f"標準時間: {standard_time} h/日（env より）")

    print("マスターデータを読み込み中...")
    period_proj             = load_period_proj_list(list_dir)
    period_inc              = load_period_inc_list(list_dir)
    item_map, hinku_group_map = load_hinku_list(list_dir)
    proj_total = sum(len(v) for v in period_proj.values())
    inc_total  = sum(len(v) for v in period_inc.values())
    print(f"  proj_list: 期数={len(period_proj)}, 計{proj_total}件 / "
          f"inc_list: 期数={len(period_inc)}, 計{inc_total}件 / "
          f"品区グループ: {len(hinku_group_map)}件")

    print("コストデータを集計中...")
    all_costs = load_all_costs(results_dir, period_proj, period_inc, item_map, hinku_group_map)
    print(f"  allCosts: {len(all_costs)} レコード")

    print("工数データを集計中...")
    all_hours = load_all_hours(results_dir, period_proj, period_inc, item_map, hinku_group_map)
    print(f"  allHours: {len(all_hours)} レコード")

    print("区分別工数を集計中...")
    kubun_hours = load_kubun_hours(results_dir, period_proj, period_inc, hinku_group_map, standard_time)
    print(f"  kubunHours: {len(kubun_hours)} レコード")

    print("人工数データを集計中...")
    monthly_person_months = load_person_months(results_dir)
    print(f"  monthlyPersonMonths: {len(monthly_person_months)} 件")

    print("インシデント件名マッピングを読み込み中...")
    inc_name_map = load_inc_name_map(list_dir, inc_xlsx)
    print(f"  incNameMap: {len(inc_name_map)} 件")

    print(f"data.js を出力中: {output_file}")
    write_data_js(output_file, all_costs, all_hours, kubun_hours, monthly_person_months, standard_time, inc_name_map)

    size_kb = os.path.getsize(output_file) // 1024
    print(f"完了 ({size_kb} KB)")


if __name__ == '__main__':
    main()
