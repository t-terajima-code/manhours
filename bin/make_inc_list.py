# -*- coding: utf-8 -*-
"""
インシデント管理表.xlsx から事業年度別の <期>_inc_list.csv を生成する。

各「◆XXX期」シートを走査し、ｲﾝｼﾃﾞﾝﾄ№/ＮＧＫｙ/品区 を抽出して
list/<期>_inc_list.csv (CP932, header: inc_num,naigaikubun,hinku) を出力する。
"""
import argparse
import csv
import os
import re
import openpyxl

DEFAULT_XLSX = "インシデント管理表.xlsx"
VALID_NAIGAI = {'N', 'G', 'K', 'y', 'Y'}


def parse_env_file(env_file_path):
    """envファイル: 1行目=root, 5行目(index 4)=list"""
    if not os.path.exists(env_file_path):
        raise FileNotFoundError(f"設定ファイル '{env_file_path}' が見つかりません。")
    with open(env_file_path, 'r', encoding='cp932') as f:
        lines = [line.strip() for line in f if line.strip()]
    if len(lines) < 5:
        raise ValueError("envファイルの行数が不足しています。")
    # env 1行目の絶対パスを優先。配布先で別PC/別ドライブに移動して 1行目が実在しない場合は、
    # env の場所(bin/)の親=パッケージルートを自動解決して動くようにする。
    _auto_root = os.path.dirname(os.path.dirname(os.path.abspath(env_file_path)))
    root_dir = lines[0] if os.path.isdir(lines[0]) else _auto_root
    list_dir = os.path.join(root_dir, lines[4].lstrip('\\/'))
    return root_dir, list_dir


def extract_period(sheet_name):
    """'◆189期' から 189 を取り出す。マッチしなければ None。"""
    m = re.match(r'^◆(\d+)期$', sheet_name.strip())
    return int(m.group(1)) if m else None


def detect_header(rows):
    """ヘッダー行を探し、(header_row_idx, {key: col_idx}) を返す。いずれも0始まり。
    rows は ws.iter_rows(values_only=True) を list 化したもの（read_only対応のため
    ランダムアクセスでなくメモリ上の2次元リストを走査する）。"""
    targets = {
        'inc_num': lambda s: 'ｲﾝｼﾃﾞﾝﾄ' in s or 'インシデント' in s,
        'naigai': lambda s: 'ＮＧＫ' in s or 'NGK' in s.upper(),
        'hinku': lambda s: s == '品区',
    }
    for row_idx in range(min(15, len(rows))):
        cols = {}
        for c, v in enumerate(rows[row_idx]):
            if v is None:
                continue
            s = str(v).strip()
            for key, matcher in targets.items():
                if key not in cols and matcher(s):
                    cols[key] = c
        if all(k in cols for k in targets):
            return row_idx, cols
    raise ValueError("ヘッダー行（ｲﾝｼﾃﾞﾝﾄ№/ＮＧＫｙ/品区）が見つかりません。")


def normalize_int_str(v):
    """セル値を「整数文字列」に正規化。float なら小数点以下を落とす。"""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def extract_data(rows, sheet_name):
    """シートから (inc_num, naigai, hinku) のリストとエラー一覧を返す。
    rows は ws.iter_rows(values_only=True) を list 化したもの。"""
    header_row, cols = detect_header(rows)
    data = []
    errors = []
    seen = {}  # inc_num -> (row_no, naigai, hinku)

    def get(row, key):
        c = cols[key]
        return row[c] if c < len(row) else None

    for row_idx in range(header_row + 1, len(rows)):
        row = rows[row_idx]
        row_no = row_idx + 1  # 表示用に1始まりへ
        inc_raw = get(row, 'inc_num')
        naigai_raw = get(row, 'naigai')
        hinku_raw = get(row, 'hinku')

        # 全カラム空 → 空行としてスキップ（エラーにしない）
        if inc_raw is None and naigai_raw is None and hinku_raw is None:
            continue

        # 各セルの空チェック
        if inc_raw is None or str(inc_raw).strip() == '':
            errors.append(f"  [{sheet_name}] 行{row_no}: インシデント№が空 (ＮＧＫｙ={naigai_raw}, 品区={hinku_raw})")
            continue
        if naigai_raw is None or str(naigai_raw).strip() == '':
            errors.append(f"  [{sheet_name}] 行{row_no}: ＮＧＫｙが空 (ｲﾝｼﾃﾞﾝﾄ№={inc_raw}, 品区={hinku_raw})")
            continue
        if hinku_raw is None or str(hinku_raw).strip() == '':
            errors.append(f"  [{sheet_name}] 行{row_no}: 品区が空 (ｲﾝｼﾃﾞﾝﾄ№={inc_raw}, ＮＧＫｙ={naigai_raw})")
            continue

        # 正規化
        inc_num = normalize_int_str(inc_raw)
        naigai = str(naigai_raw).strip()
        hinku = normalize_int_str(hinku_raw)

        # ＮＧＫｙ の値域チェック
        if naigai not in VALID_NAIGAI:
            errors.append(f"  [{sheet_name}] 行{row_no}: ＮＧＫｙ='{naigai}' は無効 (許容: N,G,K,y)")
            continue

        # 重複チェック
        if inc_num in seen:
            prev_row, prev_n, prev_h = seen[inc_num]
            if prev_n != naigai or prev_h != hinku:
                errors.append(
                    f"  [{sheet_name}] 行{row_no}: ｲﾝｼﾃﾞﾝﾄ№={inc_num} 値矛盾 "
                    f"(前回行{prev_row}: {prev_n}/{prev_h}, 今回: {naigai}/{hinku})"
                )
                continue
            # 同一値の重複は黙ってスキップ
            continue

        seen[inc_num] = (row_no, naigai, hinku)
        data.append((inc_num, naigai, hinku))

    return data, errors


def write_csv(out_path, data):
    with open(out_path, 'w', encoding='cp932', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['inc_num', 'naigaikubun', 'hinku'])
        for row in data:
            writer.writerow(row)


def main():
    parser = argparse.ArgumentParser(description='インシデント管理表から <期>_inc_list.csv を生成')
    parser.add_argument('--env', type=str, default='env')
    parser.add_argument('--period', type=int, default=None,
                        help='対象の期番号（例: 189）。未指定なら ◆XXX期 シートを全件処理')
    parser.add_argument('--xlsx', type=str, default=None,
                        help='インシデント管理表.xlsx のパス。未指定ならルート直下を使用')
    args = parser.parse_args()

    try:
        root_dir, list_dir = parse_env_file(args.env)
    except Exception as e:
        print(f"【エラー】{e}")
        return

    xlsx_path = args.xlsx or os.path.join(root_dir, DEFAULT_XLSX)
    if not os.path.exists(xlsx_path):
        print(f"【エラー】インシデント管理表が見つかりません: {xlsx_path}")
        return

    # read_only=True: 埋め込み画像の読み込みを回避し、新しい openpyxl でも
    # クラッシュせず読めるようにする（従来運用の堅牢化）
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    # 対象シート選定
    targets = []
    for sheet_name in wb.sheetnames:
        period = extract_period(sheet_name)
        if period is None:
            continue
        if args.period is not None and period != args.period:
            continue
        targets.append((sheet_name, period))

    if not targets:
        if args.period is not None:
            print(f"【エラー】◆{args.period}期 シートが見つかりません。")
        else:
            print("【エラー】◆XXX期 パターンのシートが一つも見つかりません。")
        return

    # 全シートを先に走査し、エラーが無いことを確認してから書き込む
    all_errors = []
    results = []
    for sheet_name, period in targets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        try:
            data, errors = extract_data(rows, sheet_name)
        except ValueError as e:
            print(f"【エラー】シート '{sheet_name}': {e}")
            wb.close()
            return
        all_errors.extend(errors)
        results.append((sheet_name, period, data))

    wb.close()

    if all_errors:
        print("【エラー】以下の不正データを修正してから再実行してください:")
        for err in all_errors:
            print(err)
        return

    # 書き込み（無言で上書き）
    for sheet_name, period, data in results:
        out_path = os.path.join(list_dir, f"{period}_inc_list.csv")
        write_csv(out_path, data)
        print(f"出力完了: {out_path} ({len(data)}件, シート={sheet_name})")


if __name__ == "__main__":
    main()
