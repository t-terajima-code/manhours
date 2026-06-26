# -*- coding: utf-8 -*-
r"""build_inc_master.py

インシデント案件の真のマスタである「インシデント管理表.xlsx」から、
期別の一元マスタ <期>_inc_master.csv (CP932, ヘッダー付き) を生成する。

データフロー上の位置づけ:
  インシデント管理表.xlsx
      -> <期>_inc_master.csv  (一元マスタ ... 本スクリプト)
      -> <期>_inc_list.csv    (派生; derive_inc_list.py)
      -> inc_name_list 相当    (派生; code,name)

xlsx の構造 (調査済み):
  - 期別シート名: ◆189期 / ◆188期 のように「◆+期番号+期」。
    本スクリプトは「◆+数字+期」を期番号(数字3桁)でマッチして特定する。
  - 各期シートの上部にはタイトル行・空行・注記行(例 "189001リンク")があり、
    ヘッダー行の行番号は決め打ちできない。
    ヘッダー行は「ｲﾝｼﾃﾞﾝﾄ」を含むセルと「件名」(全角空白入り "件　　　名")を
    含むセルの両方を持つ行を探して特定する。
  - 列はヘッダー名でマッピングする (期で並びが変わっても耐えるため)。
    確認された列: 備考 / ｲﾝｼﾃﾞﾝﾄ№ / 依頼日 / 製品群 / ＮＧＫｙ / 品区 /
                  件名 / 担当 / 業務内容 / 依頼の流れ / 納期 / 完了日
    ※ ヘッダー行の右側(列N以降)に別表(品区リスト等)の断片が混入することが
      あるが、必要列をヘッダー名で拾うため影響しない。
  - データ行は ｲﾝｼﾃﾞﾝﾄ№ に数値コード(例 183026)を持つ行のみ採用。
    コードが数値でない/空の行(タイトル, "189001リンク" 等の注記)はスキップ。
  - ＮＧＫｙ = 内外製区分 (N/G/K/y)。半角で格納済み。空欄の行もまれに存在する。
  - 品区 = 品区コード。内外製が y の行は品区に y10/y20 等が入る(現行 inc_list と同形式)。

出力契約:
  - list/<期>_inc_master.csv, CP932, ヘッダー付き
  - 必須列: code,name,内外製,品区
  - 追加保持列: 製品群,担当,依頼日,納期,完了日,備考,業務内容
  - code 昇順、name は前後空白除去、日付列は YYYY-MM-DD 文字列化
  - 同一コードが複数行ある場合は最初の出現を採用し、警告を表示する。

使い方 (bin から):
  python build_inc_master.py
  python build_inc_master.py --periods 188 189
  python build_inc_master.py --xlsx ..\インシデント管理表.xlsx --out-dir ..\list
"""
import argparse
import csv
import os
import re
import sys
import warnings

import openpyxl


# 出力する列順 (先頭4列が必須、以降は任意の追加列)
OUTPUT_COLUMNS = [
    'code', 'name', '内外製', '品区',
    '製品群', '担当', '依頼日', '納期', '完了日', '備考', '業務内容',
]

# xlsx ヘッダー名 -> 出力列名 の対応。
# キーは「空白(半角/全角)を除去した」ヘッダー文字列で比較する。
HEADER_ALIASES = {
    'ｲﾝｼﾃﾞﾝﾄ': 'code',        # ｲﾝｼﾃﾞﾝﾄ№ (№/Noの揺れに耐えるため前方一致扱い)
    '件名': 'name',
    'ＮＧＫｙ': '内外製',
    'NGKy': '内外製',
    '品区': '品区',
    '製品群': '製品群',
    '担当': '担当',
    '依頼日': '依頼日',
    '納期': '納期',
    '完了日': '完了日',
    '備考': '備考',
    '業務内容': '業務内容',
}

# 日付として YYYY-MM-DD 文字列化する列
DATE_COLUMNS = {'依頼日', '納期', '完了日'}

# CP932 に無い記号を、意味が近い CP932 表現に置換する。
# (業務内容・依頼の流れ欄に矢印類が混入するため。経理側ツールの文字化け回避)
_CP932_REPLACEMENTS = {
    '⇨': '→',   # ⇨ 白抜き右矢印
    '⇒': '→',   # ⇒ 二重右矢印
    '→': '→',   # → (CP932 にあるが念のため明示)
    '➔': '→',   # ➔
    '➜': '→',   # ➜
    '⮕': '→',   # ⮕
    '〜': '〜',   # 〜 波ダッシュ -> 全角チルダ
    '～': '〜',
}


def _to_cp932_safe(s):
    """CP932 で安全に書ける文字列にする。

    既知の非 CP932 記号は近い CP932 文字へ置換し、それでも表現できない
    文字は '?' にフォールバックする (黙って欠落させない)。
    """
    if s == '':
        return s
    for src, dst in _CP932_REPLACEMENTS.items():
        if src in s:
            s = s.replace(src, dst)
    # 残る非 CP932 文字は ? に (encode/decode ラウンドトリップで検出)
    return s.encode('cp932', errors='replace').decode('cp932')


def _squash(s):
    """ヘッダー比較用に半角/全角空白を除去した文字列を返す。"""
    if s is None:
        return ''
    return str(s).replace(' ', '').replace('　', '')


def find_period_sheet(wb, period):
    """ワークブックから「◆<period>期」相当のシート名を返す。無ければ None。

    シート名に含まれる数字が period と一致し、かつ短い名前(期シート)を採用する。
    末尾の集計シート群(例 "180年間(部署)") を誤って拾わないため長さ上限を設ける。
    """
    for name in wb.sheetnames:
        digits = ''.join(ch for ch in name if ch.isdigit())
        if digits == str(period) and len(name) <= 6:
            return name
    return None


def find_header_row(ws, max_scan=20):
    """ヘッダー行(1始まり)と、その行の値リストを返す。見つからなければ (None, None)。

    「ｲﾝｼﾃﾞﾝﾄ」を含むセルと「件名(=件＋名)」を含むセルの両方を持つ行を採用する。
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan,
                                         values_only=True), start=1):
        vals = [('' if c is None else str(c)) for c in row]
        squashed = [_squash(v) for v in vals]
        has_inc = any('ｲﾝｼﾃﾞﾝﾄ' in s for s in squashed)
        has_name = any(('件' in v and '名' in v) for v in vals)
        if has_inc and has_name:
            return i, vals
    return None, None


def map_columns(header_vals):
    """ヘッダー値リストから {出力列名: 列index} を返す。

    code 列は「ｲﾝｼﾃﾞﾝﾄ」を含むセル、それ以外は完全一致(空白除去後)で対応付ける。
    同名候補が複数あれば最初に現れた列を採用する。
    """
    col_index = {}
    for idx, raw in enumerate(header_vals):
        squashed = _squash(raw)
        if squashed == '':
            continue
        # code は前方一致 (№/No の揺れ・全角空白に耐える)
        if 'code' not in col_index and 'ｲﾝｼﾃﾞﾝﾄ' in squashed:
            col_index['code'] = idx
            continue
        target = HEADER_ALIASES.get(squashed)
        if target and target not in col_index:
            col_index[target] = idx
    return col_index


def normalize_code(raw):
    """ｲﾝｼﾃﾞﾝﾄ№ を整数文字列に正規化する。数値でなければ None。"""
    if raw is None:
        return None
    s = str(raw).strip()
    if s == '':
        return None
    s = s.split('.')[0]  # "183026.0" -> "183026"
    if not s.isdigit():
        return None
    return str(int(s))


def normalize_date(value):
    """日付セルを YYYY-MM-DD 文字列にする。日付でなければ前後空白除去した文字列。"""
    if value is None:
        return ''
    # openpyxl は日付セルを datetime で返す (data_only=True)
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    # "2020-04-29 00:00:00" のような文字列が来た場合は日付部分だけ残す
    m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
    if m:
        return '%04d-%02d-%02d' % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return s


def normalize_text(value):
    """一般セルを前後空白除去した文字列にする。"""
    if value is None:
        return ''
    return str(value).strip()


def extract_period(ws, header_row, col_index):
    """ヘッダー以降のデータ行から案件レコードのリストを返す。

    返り値: (records, dup_codes)
      records  : OUTPUT_COLUMNS をキーに持つ dict のリスト (code 昇順)
      dup_codes: 重複していたコード文字列のリスト (出現順)
    """
    code_col = col_index.get('code')
    if code_col is None:
        raise ValueError('ｲﾝｼﾃﾞﾝﾄ№ 列が特定できませんでした')

    records = {}     # code -> record dict
    dup_codes = []   # 重複コード

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if code_col >= len(row):
            continue
        code = normalize_code(row[code_col])
        if code is None:
            continue  # タイトル/注記/空行はスキップ

        if code in records:
            dup_codes.append(code)
            continue  # 最初の出現を優先

        rec = {c: '' for c in OUTPUT_COLUMNS}
        rec['code'] = code
        for out_name in OUTPUT_COLUMNS:
            if out_name == 'code':
                continue
            idx = col_index.get(out_name)
            if idx is None or idx >= len(row):
                continue
            raw = row[idx]
            if out_name in DATE_COLUMNS:
                rec[out_name] = normalize_date(raw)
            else:
                rec[out_name] = normalize_text(raw)
        records[code] = rec

    ordered = [records[c] for c in sorted(records.keys(), key=lambda c: int(c))]
    return ordered, dup_codes


def write_master(records, output_path):
    """レコードを CP932 / ヘッダー付き CSV に出力する。"""
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, 'w', encoding='cp932', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(OUTPUT_COLUMNS)
        for rec in records:
            writer.writerow([_to_cp932_safe(rec.get(c, '')) for c in OUTPUT_COLUMNS])
    return len(records)


def build_one_period(wb, period, out_dir):
    """1期分の inc_master を生成する。生成件数を返す。"""
    sheet = find_period_sheet(wb, period)
    if sheet is None:
        print(f"【エラー】{period}期のシートが見つかりません")
        return None
    ws = wb[sheet]
    header_row, header_vals = find_header_row(ws)
    if header_row is None:
        print(f"【エラー】{period}期({sheet}) のヘッダー行を特定できませんでした")
        return None
    col_index = map_columns(header_vals)
    missing = [c for c in ('code', 'name', '内外製', '品区')
               if c not in col_index]
    if missing:
        print(f"【エラー】{period}期({sheet}) で必須列が見つかりません: {missing}")
        return None

    records, dup_codes = extract_period(ws, header_row, col_index)

    # 内外製が空のレコードを警告 (現行 inc_list に載らない可能性がある)
    blank_naigai = [r['code'] for r in records if r['内外製'] == '']
    if blank_naigai:
        print(f"  【警告】{period}期: 内外製(ＮＧＫｙ)が空のコード "
              f"({len(blank_naigai)}件): {', '.join(blank_naigai)}")
    if dup_codes:
        print(f"  【警告】{period}期: 重複コード(最初の出現を採用) "
              f"({len(dup_codes)}件): {', '.join(dup_codes)}")

    output_path = os.path.join(out_dir, f'{period}_inc_master.csv')
    count = write_master(records, output_path)
    print(f"  {period}_inc_master.csv を生成: {output_path} ({count}件, シート={sheet})")
    return count


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='インシデント管理表.xlsx から <期>_inc_master.csv を生成する')
    bin_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(bin_dir)
    parser.add_argument(
        '--xlsx',
        default=os.path.join(project_root, 'インシデント管理表.xlsx'),
        help='入力 xlsx (既定: ../インシデント管理表.xlsx)')
    parser.add_argument(
        '--out-dir',
        default=os.path.join(project_root, 'list'),
        help='出力先ディレクトリ (既定: ../list)')
    parser.add_argument(
        '--periods', nargs='+', default=['188', '189'],
        help='生成対象の期番号 (既定: 188 189)')
    args = parser.parse_args(argv)

    if not os.path.isfile(args.xlsx):
        print(f"【エラー】xlsx が見つかりません: {args.xlsx}")
        return 1

    # Data Validation の UserWarning を抑制
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    total = 0
    ok = True
    for period in args.periods:
        count = build_one_period(wb, period, args.out_dir)
        if count is None:
            ok = False
        else:
            total += count
    wb.close()

    print(f"完了: 合計 {total} 件")
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
